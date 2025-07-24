import openpyxl

wb = openpyxl.load_workbook("SkillSync_Test_Cases.xlsx")
ws = wb["Sheet1"]

ws['A1'] = 'category'
ws['B1'] = 'test_case_id'
ws['C1'] = 'description'
ws['D1'] = 'input'
ws['E1'] = 'expected_result'

wb.save("SkillSync_Test_Cases.xlsx")
print("✅ Excel headers updated successfully.")
